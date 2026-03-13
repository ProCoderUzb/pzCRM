import io
from datetime import datetime
from django.http import HttpResponse
from django.db.models import Count, Q
from rest_framework import views, viewsets, filters, status
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
from config.permissions import IsCEOOrDev, IsAdminOrCEOOrDev
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .models import AuditLog
from students.models import Student, Lead
from academics.models import CourseClass, Attendance
from finance.models import Payment, Expense
from users.models import User


# ─── Helpers ──────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def make_workbook(headers: list, rows: list, sheet_name: str = "Data") -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 4, 14)

    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    return wb


def xlsx_response(wb: openpyxl.Workbook, filename: str) -> HttpResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


# ─── Audit Log ────────────────────────────────────────────────────────────────
class AuditLogView(views.APIView):
    permission_classes = [IsCEOOrDev]
    def get(self, request):
        qs = AuditLog.objects.select_related('user').all()

        model = request.query_params.get('model')
        action = request.query_params.get('action')
        search = request.query_params.get('search')
        limit = int(request.query_params.get('limit', 100))

        if model:
            qs = qs.filter(model_name__icontains=model)
        if action:
            qs = qs.filter(action=action)
        if search:
            qs = qs.filter(
                Q(object_repr__icontains=search) |
                Q(user__first_name__icontains=search) |
                Q(user__last_name__icontains=search) |
                Q(user__username__icontains=search)
            )

        data = []
        for log in qs[:limit]:
            data.append({
                'id': log.id,
                'timestamp': log.timestamp.strftime('%Y-%m-%d %H:%M'),
                'user': log.user.get_full_name() or log.user.username if log.user else 'System',
                'action': log.action,
                'model': log.model_name,
                'object_id': log.object_id,
                'object_repr': log.object_repr,
                'changes': log.changes,
            })
        return Response(data)


# ─── EXCEL EXPORTS ────────────────────────────────────────────────────────────
class ExportStudentsView(views.APIView):
    permission_classes = [IsAdminOrCEOOrDev]
    def get(self, request):
        qs = Student.objects.all()
        headers = ['ID', 'Full Name', 'Phone', 'Parent Name', 'Parent Phone', 'Active', 'Balance', 'Notes']
        rows = [(s.id, s.full_name, s.phone_number, s.parent_name, s.parent_phone, 'Yes' if s.is_active else 'No', float(s.balance), s.notes) for s in qs]
        wb = make_workbook(headers, rows, 'Students')
        return xlsx_response(wb, f'students_{datetime.now():%Y%m%d}.xlsx')


class ExportLeadsView(views.APIView):
    permission_classes = [IsAdminOrCEOOrDev]
    def get(self, request):
        qs = Lead.objects.all()
        headers = ['ID', 'Full Name', 'Phone', 'Status', 'Notes', 'Created']
        rows = [(l.id, l.full_name, l.phone_number, l.status, l.notes, str(l.created_at.date())) for l in qs]
        wb = make_workbook(headers, rows, 'Leads')
        return xlsx_response(wb, f'leads_{datetime.now():%Y%m%d}.xlsx')


class ExportPaymentsView(views.APIView):
    permission_classes = [IsCEOOrDev]
    def get(self, request):
        qs = Payment.objects.select_related('student').all()
        headers = ['ID', 'Student', 'Amount', 'Date', 'Method', 'Notes']
        rows = [(p.id, p.student.full_name, float(p.amount), str(p.date), p.method, p.notes) for p in qs]
        wb = make_workbook(headers, rows, 'Payments')
        return xlsx_response(wb, f'payments_{datetime.now():%Y%m%d}.xlsx')


class ExportExpensesView(views.APIView):
    permission_classes = [IsCEOOrDev]
    def get(self, request):
        qs = Expense.objects.all()
        headers = ['ID', 'Title', 'Amount', 'Date', 'Category', 'Notes']
        rows = [(e.id, e.title, float(e.amount), str(e.date), e.category, e.notes) for e in qs]
        wb = make_workbook(headers, rows, 'Expenses')
        return xlsx_response(wb, f'expenses_{datetime.now():%Y%m%d}.xlsx')


class ExportAttendanceView(views.APIView):
    permission_classes = [IsAdminOrCEOOrDev]
    def get(self, request):
        qs = Attendance.objects.select_related('student', 'course_class').all()
        date_from = request.query_params.get('from')
        date_to = request.query_params.get('to')
        class_id = request.query_params.get('class')
        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)
        if class_id:
            qs = qs.filter(course_class_id=class_id)
        headers = ['Date', 'Class', 'Student', 'Status', 'Notes']
        rows = [(str(a.date), a.course_class.name, a.student.full_name, a.status, a.notes) for a in qs]
        wb = make_workbook(headers, rows, 'Attendance')
        return xlsx_response(wb, f'attendance_{datetime.now():%Y%m%d}.xlsx')


        return xlsx_response(wb, f'attendance_{datetime.now():%Y%m%d}.xlsx')


# ─── EXCEL IMPORTS ────────────────────────────────────────────────────────────
class ImportStudentsView(views.APIView):
    permission_classes = [IsAdminOrCEOOrDev]
    parser_classes = [MultiPartParser]

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file uploaded.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            created = 0
            skipped = 0
            errors = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0] and not row[1]:
                    continue
                full_name = str(row[1] or '').strip()
                if not full_name:
                    skipped += 1
                    continue
                try:
                    Student.objects.get_or_create(
                        full_name=full_name,
                        defaults={
                            'phone_number': str(row[2] or ''),
                            'parent_name': str(row[3] or ''),
                            'parent_phone': str(row[4] or ''),
                            'is_active': str(row[5] or 'Yes').lower() == 'yes',
                            'notes': str(row[7] or ''),
                        }
                    )
                    created += 1
                except Exception as e:
                    errors.append(f"Row {full_name}: {str(e)}")
            return Response({'created': created, 'skipped': skipped, 'errors': errors})
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class ImportLeadsView(views.APIView):
    permission_classes = [IsAdminOrCEOOrDev]
    parser_classes = [MultiPartParser]

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file uploaded.'}, status=status.HTTP_400_BAD_REQUEST)
        VALID_STATUSES = ['NEW', 'CONTACTED', 'TRIAL_SCHEDULED', 'TRIAL_DONE', 'ENROLLED', 'REJECTED']
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            created = 0
            skipped = 0
            errors = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                full_name = str(row[1] or '').strip()
                if not full_name:
                    skipped += 1
                    continue
                try:
                    raw_status = str(row[3] or 'NEW').upper()
                    lead_status = raw_status if raw_status in VALID_STATUSES else 'NEW'
                    Lead.objects.get_or_create(
                        full_name=full_name,
                        defaults={
                            'phone_number': str(row[2] or ''),
                            'status': lead_status,
                            'notes': str(row[4] or ''),
                        }
                    )
                    created += 1
                except Exception as e:
                    errors.append(f"Row {full_name}: {str(e)}")
            return Response({'created': created, 'skipped': skipped, 'errors': errors})
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


# ─── ATTENDANCE STATS ─────────────────────────────────────────────────────────
class AttendanceStatsView(views.APIView):
    permission_classes = [IsAdminOrCEOOrDev]
    def get(self, request):
        class_id = request.query_params.get('class')
        date_from = request.query_params.get('from')
        date_to = request.query_params.get('to')

        qs = Attendance.objects.select_related('student', 'course_class').all()
        if class_id:
            qs = qs.filter(course_class_id=class_id)
        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)

        # Overall counts
        total = qs.count()
        by_status = list(qs.values('status').annotate(count=Count('id')).order_by('status'))

        # Per-student breakdown
        student_stats = []
        if class_id:
            for s_data in qs.values('student__full_name', 'student_id').distinct():
                sub = qs.filter(student_id=s_data['student_id'])
                row = {'student': s_data['student__full_name'], 'total': sub.count()}
                for st in ['PRESENT', 'ABSENT', 'LATE', 'EXCUSED']:
                    row[st.lower()] = sub.filter(status=st).count()
                row['rate'] = round(row['present'] / row['total'] * 100, 1) if row['total'] else 0
                student_stats.append(row)
            student_stats.sort(key=lambda x: x['rate'])

        # Per-date trend
        daily = list(qs.values('date', 'status').annotate(count=Count('id')).order_by('date'))

        return Response({
            'total_records': total,
            'by_status': by_status,
            'student_stats': student_stats,
            'daily_trend': daily,
        })
